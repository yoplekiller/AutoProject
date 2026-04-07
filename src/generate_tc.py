from jira import JIRA
from groq import Groq
from dotenv import load_dotenv
import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

# 클라이언트 초기화
jira = JIRA(
    server=os.getenv("JIRA_URL"),
    basic_auth=(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN"))
)
groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))

project_key = os.getenv("JIRA_PROJECT_KEY")

# Jira 티켓 조회
issues = jira.search_issues(f"project={project_key} ORDER BY created DESC", maxResults=5)

print(f"=== [{project_key}] TC 자동 생성 시작 ===\n")

results = []

for issue in issues:
    summary = issue.fields.summary
    status = issue.fields.status.name
    description = issue.fields.description or "설명 없음"

    print(f"[{issue.key}] {summary} (상태: {status})")
    print("-" * 50)

    # Groq에 TC 생성 요청
    response = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {
                "role": "system",
                "content": "당신은 QA 엔지니어입니다. Jira 티켓 정보를 보고 구체적인 테스트 케이스를 작성해주세요."
            },
            {
                "role": "user",
                "content": f"""다음 Jira 티켓에 대한 테스트 케이스를 기능 설명을 참고하고 테스트 포인트의 갯수대로 테스트를 작성해주세요.

티켓 제목: {summary}

티켓 설명:
{description}

반드시 아래 JSON 배열 형식으로만 응답하세요. 설명이나 마크다운 없이 JSON만 출력하세요.

[
  {{
    "tc_id": "TC-001",
    "테스트항목": "",
    "사전조건": "",
    "테스트단계": "1. 단계1\\n2. 단계2\\n3. 단계3",
    "기대결과": ""
  }}
]"""
            }
        ]
    )

    tc_raw = response.choices[0].message.content.strip()

    try:
        tc_list = json.loads(tc_raw)
    except json.JSONDecodeError:
        # JSON 파싱 실패 시 원본 텍스트 보존
        tc_list = [{"tc_id": "ERROR", "테스트항목": tc_raw, "사전조건": "", "테스트단계": "", "기대결과": ""}]

    for tc in tc_list:
        print(f"  [{tc.get('tc_id')}] {tc.get('테스트항목')}")

    print("\n")

    results.append({
        "issue_key": issue.key,
        "summary": summary,
        "status": status,
        "description": description,
        "test_cases": tc_list
    })

# reports/ 폴더에 저장
os.makedirs("reports", exist_ok=True)
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

# JSON 저장
json_filename = f"reports/tc_{timestamp}.json"
with open(json_filename, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"✅ JSON 저장 완료: {json_filename}")

# 엑셀 저장
xlsx_filename = f"reports/tc_{timestamp}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트케이스"

# 헤더 스타일
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = ["TC ID", "티켓 키", "요약", "상태", "테스트 항목", "사전 조건", "테스트 단계", "기대 결과"]
col_widths = [10, 14, 25, 12, 25, 25, 40, 30]

for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws.column_dimensions[cell.column_letter].width = width

ws.row_dimensions[1].height = 25

# 데이터 입력
row = 2
for item in results:
    for tc in item["test_cases"]:
        ws.cell(row=row, column=1, value=tc.get("tc_id", ""))
        key_cell = ws.cell(row=row, column=2, value=item["issue_key"])
        key_cell.hyperlink = f"{os.getenv('JIRA_URL')}/browse/{item['issue_key']}"
        key_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row, column=3, value=item["summary"])
        ws.cell(row=row, column=4, value=item["status"])
        ws.cell(row=row, column=5, value=tc.get("테스트항목", ""))
        ws.cell(row=row, column=6, value=tc.get("사전조건", ""))
        steps_cell = ws.cell(row=row, column=7, value=tc.get("테스트단계", ""))
        steps_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=8, value=tc.get("기대결과", ""))
        ws.row_dimensions[row].height = 60
        row += 1

wb.save(xlsx_filename)
print(f"✅ 엑셀 저장 완료: {xlsx_filename}")
