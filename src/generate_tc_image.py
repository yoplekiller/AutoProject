"""
이미지(스크린샷/PNG/JPG) 기획서 기반 TC 자동 생성
Groq Vision 모델(Llama 4)로 이미지 직접 분석
사용법: python src/generate_tc_image.py docs/스크린샷.png
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os
import json
import argparse
import base64
from datetime import datetime
from groq import Groq
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

parser = argparse.ArgumentParser()
parser.add_argument("file", help="기획서 이미지 파일 경로 (png, jpg, jpeg)")
args = parser.parse_args()

file_path = args.file
if not os.path.exists(file_path):
    print(f"[ERROR] 파일을 찾을 수 없습니다: {file_path}")
    sys.exit(1)

ext = os.path.splitext(file_path)[1].lower()
mime_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
mime_type = mime_map.get(ext)
if not mime_type:
    print(f"[ERROR] 지원하지 않는 형식입니다: {ext} (png, jpg, jpeg만 가능)")
    sys.exit(1)

doc_name = os.path.splitext(os.path.basename(file_path))[0]
print(f"=== [{doc_name}] TC 자동 생성 시작 ===\n")
print(f"파일: {file_path}")
print(f"Groq Vision 모델로 이미지 분석 중...\n")

# 이미지 base64 인코딩
with open(file_path, "rb") as f:
    image_data = base64.b64encode(f.read()).decode("utf-8")

response = groq_client.chat.completions.create(
    model="meta-llama/llama-4-scout-17b-16e-instruct",
    messages=[
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{mime_type};base64,{image_data}"
                    }
                },
                {
                    "type": "text",
                    "text": """이 이미지는 서비스 기획서 또는 화면 설계서입니다.
이미지 내용을 분석하여 QA 테스트 케이스를 작성해주세요.

반드시 아래 JSON 배열 형식으로만 응답하세요. 설명이나 마크다운 없이 JSON만 출력하세요.

[
  {
    "tc_id": "TC-001",
    "테스트항목": "",
    "사전조건": "",
    "테스트단계": "1. 단계1\\n2. 단계2\\n3. 단계3",
    "기대결과": ""
  }
]"""
                }
            ]
        }
    ]
)

tc_raw = response.choices[0].message.content.strip()
if tc_raw.startswith("```"):
    tc_raw = "\n".join(tc_raw.split("\n")[1:])
if tc_raw.endswith("```"):
    tc_raw = tc_raw.rsplit("```", 1)[0]

try:
    tc_list = json.loads(tc_raw)
except json.JSONDecodeError:
    tc_list = [{"tc_id": "ERROR", "테스트항목": tc_raw, "사전조건": "", "테스트단계": "", "기대결과": ""}]

print(f"TC {len(tc_list)}개 생성 완료:\n")
for tc in tc_list:
    print(f"  [{tc.get('tc_id')}] {tc.get('테스트항목')}")

results = [{
    "issue_key": f"DOC-{doc_name}",
    "summary": doc_name,
    "status": "기획서",
    "description": f"이미지 기반 TC 생성: {file_path}",
    "test_cases": tc_list
}]

# 저장
os.makedirs(os.path.join(ROOT_DIR, "reports"), exist_ok=True)
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

json_filename = os.path.join(ROOT_DIR, f"reports/tc_{timestamp}.json")
with open(json_filename, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"\n[OK] JSON 저장: {json_filename}")

xlsx_filename = os.path.join(ROOT_DIR, f"reports/tc_{timestamp}.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트케이스"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
headers = ["문서명", "TC ID", "테스트 항목", "사전 조건", "테스트 단계", "기대 결과"]
col_widths = [20, 10, 30, 25, 40, 30]

for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws.column_dimensions[cell.column_letter].width = width
ws.row_dimensions[1].height = 25

row = 2
for item in results:
    for tc in item["test_cases"]:
        ws.cell(row=row, column=1, value=item["summary"])
        ws.cell(row=row, column=2, value=tc.get("tc_id", ""))
        ws.cell(row=row, column=3, value=tc.get("테스트항목", ""))
        ws.cell(row=row, column=4, value=tc.get("사전조건", ""))
        steps_cell = ws.cell(row=row, column=5, value=tc.get("테스트단계", ""))
        steps_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=6, value=tc.get("기대결과", ""))
        ws.row_dimensions[row].height = 60
        row += 1

wb.save(xlsx_filename)
print(f"[OK] 엑셀 저장: {xlsx_filename}")
print(f"\n=== 완료: TC {len(tc_list)}개 생성 ===")
