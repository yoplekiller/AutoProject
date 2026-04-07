"""
티켓 URL을 입력받아 매뉴얼 TC를 생성하는 스크립트

사용법:
  # 단일 티켓 (URL 또는 키)
  python src/generate_tc_from_url.py PROJ-123
  python src/generate_tc_from_url.py https://yourcompany.atlassian.net/browse/PROJ-123

  # 엑셀 일괄 처리 (A열에 티켓 URL/키 목록)
  python src/generate_tc_from_url.py tickets.xlsx

  # 구글 스프레드시트 일괄 처리 (A열에 티켓 URL/키 목록)
  python src/generate_tc_from_url.py https://docs.google.com/spreadsheets/d/SHEET_ID/edit

  # 입력용 템플릿 엑셀 생성
  python src/generate_tc_from_url.py --template
"""

import sys
import re
import os
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from jira import JIRA
from groq import Groq
from dotenv import load_dotenv

load_dotenv()

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CREDS_PATH = os.path.join(ROOT_DIR, os.getenv("GOOGLE_CREDENTIALS_PATH", "credentials.json"))
JIRA_URL = os.getenv("JIRA_URL", "")


# ── Jira ─────────────────────────────────────────────────────────────

def extract_issue_key(input_str: str) -> str:
    """URL 또는 이슈 키에서 Jira 이슈 키를 추출합니다."""
    url_match = re.search(r"/browse/([A-Z][A-Z0-9_]+-\d+)", input_str)
    if url_match:
        return url_match.group(1)

    key_match = re.fullmatch(r"[A-Z][A-Z0-9_]+-\d+", input_str.strip())
    if key_match:
        return input_str.strip()

    raise ValueError(
        f"유효한 Jira 티켓 URL 또는 이슈 키를 입력해주세요.\n"
        f"  예) https://yourcompany.atlassian.net/browse/PROJ-123\n"
        f"  예) PROJ-123\n"
        f"  입력값: {input_str}"
    )


def fetch_issue(jira: JIRA, issue_key: str) -> dict:
    """Jira 이슈 정보를 가져옵니다."""
    issue = jira.issue(issue_key)
    return {
        "key": issue.key,
        "summary": issue.fields.summary,
        "status": issue.fields.status.name,
        "description": issue.fields.description or "설명 없음",
        "issue_type": issue.fields.issuetype.name,
    }


# ── Groq ─────────────────────────────────────────────────────────────

def generate_test_cases(groq_client: Groq, issue: dict) -> list:
    """Groq API를 사용해 테스트 케이스를 생성합니다."""
    response = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {
                "role": "system",
                "content": (
                    "당신은 경력 있는 QA 엔지니어입니다. "
                    "Jira 티켓 정보를 바탕으로 매뉴얼 테스트 케이스를 작성합니다. "
                    "정상 흐름(Happy Path), 예외 처리, 경계값 등을 고려하여 작성하세요."
                ),
            },
            {
                "role": "user",
                "content": f"""다음 Jira 티켓에 대한 매뉴얼 테스트 케이스를 작성해주세요.

티켓 키: {issue['key']}
티켓 유형: {issue['issue_type']}
티켓 제목: {issue['summary']}

티켓 설명:
{issue['description']}

반드시 아래 JSON 배열 형식으로만 응답하세요. 설명이나 마크다운 코드블록 없이 JSON만 출력하세요.

[
  {{
    "tc_id": "TC-001",
    "테스트항목": "",
    "사전조건": "",
    "테스트단계": "1. 단계1\\n2. 단계2\\n3. 단계3",
    "기대결과": "",
    "우선순위": "High/Medium/Low"
  }}
]""",
            },
        ],
    )

    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        print(f"  [경고] JSON 파싱 실패, 원본 텍스트를 저장합니다.")
        return [
            {
                "tc_id": "TC-ERROR",
                "테스트항목": raw,
                "사전조건": "",
                "테스트단계": "",
                "기대결과": "",
                "우선순위": "",
            }
        ]


# ── 엑셀 입력/출력 ────────────────────────────────────────────────────

def create_template(output_path: str = "tickets_template.xlsx"):
    """입력용 티켓 목록 템플릿 엑셀을 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "티켓목록"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="티켓 URL 또는 이슈 키").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=2, value="메모 (선택)").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[1].height = 22

    examples = [
        ("MKQA-1", "로그인 기능 TC"),
        ("MKQA-2", "회원가입 TC"),
        (f"{JIRA_URL}/browse/MKQA-3", "URL 형식 예시"),
    ]
    for r, (key, memo) in enumerate(examples, start=2):
        ws.cell(row=r, column=1, value=key).font = Font(color="808080", italic=True)
        ws.cell(row=r, column=2, value=memo).font = Font(color="808080", italic=True)

    wb.save(output_path)
    return output_path


def read_keys_from_excel(file_path: str) -> list:
    """엑셀 A열에서 티켓 URL/키 목록을 읽어옵니다. (헤더 제외, 빈 셀 스킵)"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    keys = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            keys.append(str(val).strip())
    return keys


def save_excel(results: list, output_path: str):
    """결과를 엑셀 파일로 저장합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "테스트케이스"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    top_align = Alignment(vertical="top", wrap_text=True)

    headers = ["티켓 키", "요약", "상태", "우선순위", "TC ID", "테스트 항목", "사전 조건", "테스트 단계", "기대 결과"]
    col_widths = [12, 28, 12, 10, 10, 28, 28, 45, 35]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 25

    row = 2
    for item in results:
        for tc in item["test_cases"]:
            ws.cell(row=row, column=1, value=item["key"]).alignment = center_align
            ws.cell(row=row, column=2, value=item["summary"]).alignment = top_align
            ws.cell(row=row, column=3, value=item["status"]).alignment = center_align
            ws.cell(row=row, column=4, value=tc.get("우선순위", "")).alignment = center_align
            ws.cell(row=row, column=5, value=tc.get("tc_id", "")).alignment = center_align
            ws.cell(row=row, column=6, value=tc.get("테스트항목", "")).alignment = top_align
            ws.cell(row=row, column=7, value=tc.get("사전조건", "")).alignment = top_align
            ws.cell(row=row, column=8, value=tc.get("테스트단계", "")).alignment = top_align
            ws.cell(row=row, column=9, value=tc.get("기대결과", "")).alignment = top_align
            ws.row_dimensions[row].height = 70
            row += 1

    wb.save(output_path)


# ── 구글 스프레드시트 입력/출력 ──────────────────────────────────────

def _get_gspread_client():
    """gspread 클라이언트를 반환합니다."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("[오류] gspread 또는 google-auth 패키지가 없습니다.")
        print("  pip install gspread google-auth")
        sys.exit(1)

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(CREDS_PATH, scopes=scopes)
    return gspread.authorize(creds)


def extract_sheet_id(url: str) -> str:
    """구글 스프레드시트 URL에서 spreadsheet ID를 추출합니다."""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    # ID 직접 입력된 경우 (URL이 아닌 순수 ID)
    if re.fullmatch(r"[a-zA-Z0-9_-]{30,}", url):
        return url
    raise ValueError(f"유효한 구글 스프레드시트 URL이 아닙니다: {url}")


def read_keys_from_sheets(sheet_id: str) -> list:
    """구글 스프레드시트 첫 번째 시트 A열에서 티켓 URL/키를 읽습니다. (헤더 제외)"""
    gc = _get_gspread_client()
    sh = gc.open_by_key(sheet_id)
    ws = sh.get_worksheet(0)
    all_values = ws.col_values(1)  # A열 전체
    # 헤더(1행) 제외, 빈 값 스킵
    return [v.strip() for v in all_values[1:] if v and v.strip()]


def save_to_sheets(results: list, sheet_id: str):
    """생성된 TC를 구글 스프레드시트의 '매뉴얼 TC' 시트에 저장합니다."""
    import gspread

    gc = _get_gspread_client()
    sh = gc.open_by_key(sheet_id)

    sheet_title = "매뉴얼 TC"
    try:
        ws = sh.worksheet(sheet_title)
        ws.clear()
        print(f"  기존 '{sheet_title}' 시트를 초기화했습니다.")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_title, rows=500, cols=9)
        print(f"  '{sheet_title}' 시트를 새로 생성했습니다.")

    # 헤더
    headers = ["티켓 키", "요약", "상태", "우선순위", "TC ID", "테스트 항목", "사전 조건", "테스트 단계", "기대 결과"]
    rows = [headers]

    for item in results:
        for tc in item["test_cases"]:
            rows.append([
                item["key"],
                item["summary"],
                item["status"],
                tc.get("우선순위", ""),
                tc.get("tc_id", ""),
                tc.get("테스트항목", ""),
                tc.get("사전조건", ""),
                tc.get("테스트단계", ""),
                tc.get("기대결과", ""),
            ])

    ws.update(rows, value_input_option="RAW")

    # 헤더 스타일
    ws.format("A1:I1", {
        "backgroundColor": {"red": 0.267, "green": 0.447, "blue": 0.769},
        "textFormat": {
            "bold": True,
            "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
        },
        "horizontalAlignment": "CENTER",
    })

    # 우선순위 컬럼 색상 (D열)
    priority_colors = {
        "High":   {"red": 1.0,  "green": 0.8,  "blue": 0.8},
        "Medium": {"red": 1.0,  "green": 0.95, "blue": 0.8},
        "Low":    {"red": 0.85, "green": 0.92, "blue": 0.85},
    }
    for i, row in enumerate(rows[1:], start=2):
        priority = row[3] if len(row) > 3 else ""
        color = priority_colors.get(priority)
        if color:
            ws.format(f"D{i}", {"backgroundColor": color})

    total_tc = len(rows) - 1
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
    print(f"  구글 시트 저장 완료: {total_tc}개 TC → '{sheet_title}' 시트")
    print(f"  {sheet_url}")
    return sheet_url


# ── 처리 공통 로직 ────────────────────────────────────────────────────

def process_keys(jira: JIRA, groq_client: Groq, issue_keys: list) -> list:
    """티켓 키 목록을 순서대로 처리하여 결과를 반환합니다."""
    results = []
    total = len(issue_keys)

    for idx, key in enumerate(issue_keys, start=1):
        print(f"\n[{idx}/{total}] {key} 처리 중...")

        try:
            issue = fetch_issue(jira, key)
        except Exception as e:
            print(f"  [건너뜀] 티켓 조회 실패: {e}")
            continue

        print(f"  제목: {issue['summary']} | 상태: {issue['status']}")
        print(f"  TC 생성 중...")
        tc_list = generate_test_cases(groq_client, issue)
        print(f"  생성된 TC: {len(tc_list)}개")
        for tc in tc_list:
            print(f"    [{tc.get('tc_id')}] [{tc.get('우선순위', '-')}] {tc.get('테스트항목')}")

        results.append({
            "key": issue["key"],
            "summary": issue["summary"],
            "status": issue["status"],
            "test_cases": tc_list,
        })

    return results


# ── 메인 ─────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("사용법:")
        print("  python src/generate_tc_from_url.py MKQA-1")
        print("  python src/generate_tc_from_url.py https://company.atlassian.net/browse/MKQA-1")
        print("  python src/generate_tc_from_url.py tickets.xlsx")
        print("  python src/generate_tc_from_url.py https://docs.google.com/spreadsheets/d/SHEET_ID/edit")
        print("  python src/generate_tc_from_url.py --template")
        sys.exit(1)

    input_str = sys.argv[1]

    # ── 템플릿 생성 모드
    if input_str == "--template":
        path = create_template("tickets_template.xlsx")
        print(f"템플릿 생성 완료: {path}")
        print("A열에 티켓 URL 또는 이슈 키를 입력한 후 실행하세요.")
        return

    # ── 구글 스프레드시트 모드
    if "docs.google.com/spreadsheets" in input_str or re.fullmatch(r"[a-zA-Z0-9_-]{44}", input_str.strip()):
        try:
            sheet_id = extract_sheet_id(input_str)
        except ValueError as e:
            print(f"[오류] {e}")
            sys.exit(1)

        print(f"\n=== 구글 스프레드시트에서 티켓 목록 읽는 중... ===")
        raw_keys = read_keys_from_sheets(sheet_id)
        if not raw_keys:
            print("[오류] 스프레드시트 A열(2행~)에 티켓 URL/키가 없습니다.")
            sys.exit(1)

        issue_keys = []
        for raw in raw_keys:
            try:
                issue_keys.append(extract_issue_key(raw))
            except ValueError:
                print(f"  [건너뜀] 유효하지 않은 값: {raw}")

        if not issue_keys:
            print("[오류] 유효한 이슈 키가 없습니다.")
            sys.exit(1)

        print(f"  읽어온 티켓: {len(issue_keys)}개 → {', '.join(issue_keys)}")
        print(f"\n=== TC 생성 시작 ===")
        label = f"sheets_{sheet_id[:8]}"
        use_sheets = True

    # ── 엑셀 파일 모드
    elif input_str.endswith(".xlsx") or input_str.endswith(".xls"):
        if not os.path.exists(input_str):
            print(f"[오류] 파일을 찾을 수 없습니다: {input_str}")
            sys.exit(1)

        raw_keys = read_keys_from_excel(input_str)
        if not raw_keys:
            print("[오류] 엑셀 A열에 티켓 URL/키가 없습니다.")
            sys.exit(1)

        issue_keys = []
        for raw in raw_keys:
            try:
                issue_keys.append(extract_issue_key(raw))
            except ValueError:
                print(f"  [건너뜀] 유효하지 않은 값: {raw}")

        if not issue_keys:
            print("[오류] 유효한 이슈 키가 없습니다.")
            sys.exit(1)

        print(f"\n=== 엑셀 일괄 TC 생성 시작: {len(issue_keys)}개 티켓 ===")
        label = f"batch_{os.path.splitext(os.path.basename(input_str))[0]}"
        sheet_id = None
        use_sheets = False

    # ── 단일 티켓 모드
    else:
        try:
            issue_key = extract_issue_key(input_str)
        except ValueError as e:
            print(f"[오류] {e}")
            sys.exit(1)
        issue_keys = [issue_key]
        print(f"\n=== 티켓 TC 자동 생성 시작: {issue_key} ===")
        label = issue_key.replace("/", "_")
        sheet_id = None
        use_sheets = False

    # 클라이언트 초기화
    jira = JIRA(
        server=os.getenv("JIRA_URL"),
        basic_auth=(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN")),
    )
    groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))

    # TC 생성
    results = process_keys(jira, groq_client, issue_keys)

    if not results:
        print("\n[오류] 생성된 TC가 없습니다.")
        sys.exit(1)

    # 결과 저장
    print(f"\n=== 결과 저장 중... ===")
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"reports/tc_{label}_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  JSON 저장 완료: {json_path}")

    xlsx_path = f"reports/tc_{label}_{timestamp}.xlsx"
    save_excel(results, xlsx_path)
    print(f"  엑셀 저장 완료: {xlsx_path}")

    # 구글 시트 모드일 경우 시트에도 저장
    if use_sheets:
        save_to_sheets(results, sheet_id)

    total_tc = sum(len(r["test_cases"]) for r in results)
    print(f"\n=== 완료: {len(results)}개 티켓 / {total_tc}개 TC 생성 ===")


if __name__ == "__main__":
    main()
